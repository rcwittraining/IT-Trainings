#!/usr/bin/env python3
# RCW-NixPerm v1.0.0 — Linux File & Folder Permission Manager (GUI Edition)
# Pure GUI tool for managing Linux permissions over SSH - no commands visible
import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox
import paramiko
import os, sys, datetime, threading, json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    XLSX_OK = True
except Exception:
    XLSX_OK = False

APP_NAME = "RCW-NixPerm"
APP_VER = "1.0.0"
APP_TAG = "Linux Permission Manager"

class SSHClient:
    def __init__(self):
        self.client = None
        self.ip = None

    def connect(self, host, port, user, auth, secret, timeout=15):
        self.client = paramiko.SSHClient()
        self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        if auth == "key":
            pkey = None
            for key_cls in (paramiko.Ed25519Key, paramiko.RSAKey, paramiko.ECDSAKey, paramiko.DSSKey):
                try:
                    pkey = key_cls.from_private_key_file(secret)
                    break
                except Exception:
                    continue
            if pkey is None:
                raise ValueError("Could not load private key file.")
            self.client.connect(hostname=host, port=port, username=user,
                                pkey=pkey, timeout=timeout, look_for_keys=False, allow_agent=False)
        else:
            self.client.connect(hostname=host, port=port, username=user,
                                password=secret, timeout=timeout, look_for_keys=False, allow_agent=False)
        self.ip = self.client.get_transport().getpeername()[0]
        return True

    def run(self, command, sudo_pw=None, timeout=120):
        if not self.client:
            return ("", "not connected", -1)
        try:
            if sudo_pw:
                full = 'sudo -S -p "" %s' % command
                stdin, stdout, stderr = self.client.exec_command(full, timeout=timeout)
                try:
                    stdin.write(sudo_pw + "\n")
                    stdin.flush()
                except Exception:
                    pass
            else:
                stdin, stdout, stderr = self.client.exec_command(command, timeout=timeout)
            out = stdout.read().decode("utf-8", "replace")
            err = stderr.read().decode("utf-8", "replace")
            try:
                code = stdout.channel.recv_exit_status()
            except Exception:
                code = 0
            return (out, err, code)
        except Exception as e:
            return ("", str(e), -1)

    def close(self):
        try:
            self.client.close()
        except Exception:
            pass

def parse_ls(line):
    parts = line.split(None, 8)
    if len(parts) < 9:
        return None
    perms = parts[0]
    if len(perms) < 10 or perms[0] not in "dlbcps-":
        return None
    return {
        "perms": perms,
        "type": {"d": "Folder", "l": "Link", "-": "File"}.get(perms[0], "Other"),
        "owner": parts[2],
        "group": parts[3],
        "size": parts[4],
        "name": parts[8].split(" -> ")[0],
        "suid": perms[3] in ("s", "S"),
        "sgid": perms[6] in ("s", "S"),
        "sticky": perms[9] in ("t", "T"),
    }

def export_excel(rows, meta, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Permissions"
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="101C30", end_color="101C30", fill_type="solid")
    
    ws.merge_cells("A1:H1")
    ws["A1"].value = "%s v%s — Permission Report" % (APP_NAME, APP_VER)
    ws["A1"].font = Font(bold=True, size=14, color="101C30")
    ws["A2"] = "Host:"; ws["B2"] = meta.get("host", "")
    ws["A3"] = "Path:"; ws["B3"] = meta.get("path", "")
    ws["A4"] = "Date:"; ws["B4"] = meta.get("date", "")
    ws["A5"] = "Items:"; ws["B5"] = len(rows)
    
    headers = ["Name", "Type", "Owner", "Group", "Read", "Write", "Execute", "Special"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    
    for ri, row in enumerate(rows, 8):
        perms = row.get("perms", "")
        ws.cell(row=ri, column=1, value=row.get("name", ""))
        ws.cell(row=ri, column=2, value=row.get("type", ""))
        ws.cell(row=ri, column=3, value=row.get("owner", ""))
        ws.cell(row=ri, column=4, value=row.get("group", ""))
        ws.cell(row=ri, column=5, value="✓" if len(perms) > 1 and perms[1] == "r" else "")
        ws.cell(row=ri, column=6, value="✓" if len(perms) > 2 and perms[2] == "w" else "")
        ws.cell(row=ri, column=7, value="✓" if len(perms) > 3 and perms[3] in "xs" else "")
        special = []
        if row.get("suid"): special.append("SUID")
        if row.get("sgid"): special.append("SGID")
        if row.get("sticky"): special.append("Sticky")
        ws.cell(row=ri, column=8, value=", ".join(special) if special else "")
    
    for ci, w in enumerate([40, 10, 15, 15, 8, 8, 10, 20], 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    wb.save(path)

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("%s v%s — %s" % (APP_NAME, APP_VER, APP_TAG))
        self.geometry("1200x800")
        self.ssh = None
        self.sudo_pw = ""
        self.audit_rows = []
        self.build_ui()

    def build_ui(self):
        pad = {"padx": 8, "pady": 5}

        # Connection frame
        conn = ttk.LabelFrame(self, text="🔌 Connect to Linux Server")
        conn.pack(fill="x", **pad)
        
        ttk.Label(conn, text="Server:").grid(row=0, column=0, sticky="e", padx=5)
        self.e_host = ttk.Entry(conn, width=20); self.e_host.grid(row=0, column=1, **pad)
        self.e_host.insert(0, "192.168.1.10")
        ttk.Label(conn, text="Port:").grid(row=0, column=2, sticky="e")
        self.e_port = ttk.Entry(conn, width=6); self.e_port.grid(row=0, column=3, **pad)
        self.e_port.insert(0, "22")
        
        ttk.Label(conn, text="Username:").grid(row=1, column=0, sticky="e", padx=5)
        self.e_user = ttk.Entry(conn, width=20); self.e_user.grid(row=1, column=1, **pad)
        self.e_user.insert(0, "root")
        ttk.Label(conn, text="Password:").grid(row=1, column=2, sticky="e")
        self.e_pass = ttk.Entry(conn, width=20, show="*"); self.e_pass.grid(row=1, column=3, **pad)
        
        self.btn_connect = ttk.Button(conn, text="🔗 Connect", command=self._connect)
        self.btn_connect.grid(row=2, column=1, **pad)
        self.lbl_status = ttk.Label(conn, text="Not connected", foreground="gray")
        self.lbl_status.grid(row=2, column=2, columnspan=2, sticky="w")

        # Audit frame
        audit = ttk.LabelFrame(self, text="📊 Audit Permissions")
        audit.pack(fill="x", **pad)
        
        ttk.Label(audit, text="Folder path:").grid(row=0, column=0, sticky="e", padx=5)
        self.e_path = ttk.Entry(audit, width=40); self.e_path.grid(row=0, column=1, **pad)
        self.e_path.insert(0, "/var/www")
        self.chk_recursive = tk.BooleanVar(value=True)
        ttk.Checkbutton(audit, text="Include subfolders", variable=self.chk_recursive).grid(row=0, column=2, **pad)
        
        self.btn_audit = ttk.Button(audit, text="🔍 Scan", command=self._audit, state="disabled")
        self.btn_audit.grid(row=0, column=3, **pad)
        self.btn_export = ttk.Button(audit, text="📄 Export to Excel", command=self._export, state="disabled")
        self.btn_export.grid(row=0, column=4, **pad)

        # File list
        list_frame = ttk.LabelFrame(self, text="📁 Files and Folders")
        list_frame.pack(fill="both", expand=True, **pad)
        
        cols = ("name", "type", "owner", "group", "permissions")
        self.tree = ttk.Treeview(list_frame, columns=cols, show="headings", height=15)
        self.tree.heading("name", text="Name")
        self.tree.heading("type", text="Type")
        self.tree.heading("owner", text="Owner")
        self.tree.heading("group", text="Group")
        self.tree.heading("permissions", text="Permissions")
        self.tree.column("name", width=250)
        self.tree.column("type", width=80)
        self.tree.column("owner", width=120)
        self.tree.column("group", width=120)
        self.tree.column("permissions", width=150)
        
        sb = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        # Modify frame
        mod = ttk.LabelFrame(self, text="✏️ Modify Permissions")
        mod.pack(fill="x", **pad)
        
        # Owner/Group
        ttk.Label(mod, text="Owner:").grid(row=0, column=0, sticky="e", padx=5)
        self.e_owner = ttk.Entry(mod, width=15); self.e_owner.grid(row=0, column=1, **pad)
        ttk.Label(mod, text="Group:").grid(row=0, column=2, sticky="e")
        self.e_group = ttk.Entry(mod, width=15); self.e_group.grid(row=0, column=3, **pad)
        
        # Permission checkboxes
        ttk.Label(mod, text="Permissions:").grid(row=1, column=0, sticky="ne", padx=5, pady=5)
        perm_frame = ttk.Frame(mod)
        perm_frame.grid(row=1, column=1, columnspan=3, sticky="w", **pad)
        
        ttk.Label(perm_frame, text="Owner:", font=("", 9, "bold")).grid(row=0, column=0, sticky="w", padx=5)
        self.chk_or = tk.BooleanVar(); ttk.Checkbutton(perm_frame, text="Read", variable=self.chk_or).grid(row=0, column=1)
        self.chk_ow = tk.BooleanVar(); ttk.Checkbutton(perm_frame, text="Write", variable=self.chk_ow).grid(row=0, column=2)
        self.chk_ox = tk.BooleanVar(); ttk.Checkbutton(perm_frame, text="Execute", variable=self.chk_ox).grid(row=0, column=3)
        
        ttk.Label(perm_frame, text="Group:", font=("", 9, "bold")).grid(row=1, column=0, sticky="w", padx=5)
        self.chk_gr = tk.BooleanVar(); ttk.Checkbutton(perm_frame, text="Read", variable=self.chk_gr).grid(row=1, column=1)
        self.chk_gw = tk.BooleanVar(); ttk.Checkbutton(perm_frame, text="Write", variable=self.chk_gw).grid(row=1, column=2)
        self.chk_gx = tk.BooleanVar(); ttk.Checkbutton(perm_frame, text="Execute", variable=self.chk_gx).grid(row=1, column=3)
        
        ttk.Label(perm_frame, text="Others:", font=("", 9, "bold")).grid(row=2, column=0, sticky="w", padx=5)
        self.chk_otr = tk.BooleanVar(); ttk.Checkbutton(perm_frame, text="Read", variable=self.chk_otr).grid(row=2, column=1)
        self.chk_otw = tk.BooleanVar(); ttk.Checkbutton(perm_frame, text="Write", variable=self.chk_otw).grid(row=2, column=2)
        self.chk_otx = tk.BooleanVar(); ttk.Checkbutton(perm_frame, text="Execute", variable=self.chk_otx).grid(row=2, column=3)
        
        # Special bits
        ttk.Label(mod, text="Special:").grid(row=2, column=0, sticky="e", padx=5)
        special_frame = ttk.Frame(mod)
        special_frame.grid(row=2, column=1, columnspan=3, sticky="w", **pad)
        self.chk_suid = tk.BooleanVar(); ttk.Checkbutton(special_frame, text="SUID (run as owner)", variable=self.chk_suid).pack(side="left", padx=5)
        self.chk_sgid = tk.BooleanVar(); ttk.Checkbutton(special_frame, text="SGID (inherit group)", variable=self.chk_sgid).pack(side="left", padx=5)
        self.chk_sticky = tk.BooleanVar(); ttk.Checkbutton(special_frame, text="Sticky (restrict delete)", variable=self.chk_sticky).pack(side="left", padx=5)
        
        # ACL
        ttk.Label(mod, text="ACL User:").grid(row=3, column=0, sticky="e", padx=5)
        self.e_acl_user = ttk.Entry(mod, width=15); self.e_acl_user.grid(row=3, column=1, **pad)
        ttk.Label(mod, text="ACL Permissions:").grid(row=3, column=2, sticky="e")
        acl_frame = ttk.Frame(mod)
        acl_frame.grid(row=3, column=3, sticky="w", **pad)
        self.chk_acl_r = tk.BooleanVar(); ttk.Checkbutton(acl_frame, text="R", variable=self.chk_acl_r).pack(side="left")
        self.chk_acl_w = tk.BooleanVar(); ttk.Checkbutton(acl_frame, text="W", variable=self.chk_acl_w).pack(side="left")
        self.chk_acl_x = tk.BooleanVar(); ttk.Checkbutton(acl_frame, text="X", variable=self.chk_acl_x).pack(side="left")
        
        # Apply options
        opt_frame = ttk.Frame(mod)
        opt_frame.grid(row=4, column=0, columnspan=4, sticky="w", **pad)
        self.chk_apply_recursive = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt_frame, text="Apply to subfolders", variable=self.chk_apply_recursive).pack(side="left", padx=5)
        
        btn_frame = ttk.Frame(mod)
        btn_frame.grid(row=5, column=0, columnspan=4, sticky="e", **pad)
        self.btn_preview = ttk.Button(btn_frame, text="👁️ Preview Changes", command=self._preview, state="disabled")
        self.btn_preview.pack(side="left", padx=5)
        self.btn_apply = ttk.Button(btn_frame, text="✅ Apply Changes", command=self._apply, state="disabled")
        self.btn_apply.pack(side="left", padx=5)

        # Log
        self.log = scrolledtext.ScrolledText(self, height=6, font=("Consolas", 9))
        self.log.pack(fill="x", **pad)
        self._log("%s v%s ready.\n" % (APP_NAME, APP_VER))

    def _log(self, msg):
        self.log.insert("end", msg)
        self.log.see("end")

    def _on_select(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        item = self.tree.item(sel[0])
        vals = item["values"]
        if len(vals) >= 5:
            self.e_owner.delete(0, "end")
            self.e_owner.insert(0, vals[2])
            self.e_group.delete(0, "end")
            self.e_group.insert(0, vals[3])
            # Parse permissions from the string
            perm_str = vals[4]
            if len(perm_str) >= 10:
                self.chk_or.set(perm_str[1] == "r")
                self.chk_ow.set(perm_str[2] == "w")
                self.chk_ox.set(perm_str[3] in "xs")
                self.chk_gr.set(perm_str[4] == "r")
                self.chk_gw.set(perm_str[5] == "w")
                self.chk_gx.set(perm_str[6] in "xs")
                self.chk_otr.set(perm_str[7] == "r")
                self.chk_otw.set(perm_str[8] == "w")
                self.chk_otx.set(perm_str[9] in "xt")
                self.chk_suid.set(perm_str[3] in "sS")
                self.chk_sgid.set(perm_str[6] in "sS")
                self.chk_sticky.set(perm_str[9] in "tT")

    def _connect(self):
        self.lbl_status.config(text="Connecting...", foreground="orange")
        self.btn_connect.config(state="disabled")
        threading.Thread(target=self._do_connect, daemon=True).start()

    def _do_connect(self):
        try:
            s = SSHClient()
            s.connect(self.e_host.get(), int(self.e_port.get() or 22),
                      self.e_user.get(), "password", self.e_pass.get())
            self.ssh = s
            out, _, _ = s.run("uname -a")
            self.after(0, lambda: self._connected(out.strip().split()[0] if out else "Linux"))
        except Exception as e:
            self.after(0, lambda: self._connect_fail(str(e)))

    def _connected(self, os_name):
        self.lbl_status.config(text="✓ Connected (%s)" % os_name, foreground="green")
        self.btn_connect.config(state="disabled")
        self.btn_audit.config(state="normal")
        self.btn_apply.config(state="normal")
        self.btn_preview.config(state="normal")
        self._log("Connected to %s\n" % self.e_host.get())

    def _connect_fail(self, err):
        self.lbl_status.config(text="✗ Failed", foreground="red")
        self.btn_connect.config(state="normal")
        self._log("Error: %s\n" % err)
        messagebox.showerror("Connection Failed", err)

    def _audit(self):
        self.btn_audit.config(state="disabled")
        threading.Thread(target=self._do_audit, daemon=True).start()

    def _do_audit(self):
        path = self.e_path.get().strip()
        recursive = self.chk_recursive.get()
        cmd = 'find %s -maxdepth 10 -exec ls -ld {} + 2>/dev/null' % repr(path) if recursive else 'ls -la %s 2>/dev/null' % repr(path)
        out, _, _ = self.ssh.run(cmd, timeout=180)
        rows = []
        for line in out.splitlines():
            parsed = parse_ls(line)
            if parsed:
                rows.append(parsed)
        self.audit_rows = rows
        self.after(0, lambda: self._fill_tree(rows, path))

    def _fill_tree(self, rows, path):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for r in rows:
            perms = r["perms"]
            perm_display = "%s%s%s %s%s%s %s%s%s" % (
                "r" if perms[1] == "r" else "-",
                "w" if perms[2] == "w" else "-",
                "x" if perms[3] in "xs" else "-",
                "r" if perms[4] == "r" else "-",
                "w" if perms[5] == "w" else "-",
                "x" if perms[6] in "xs" else "-",
                "r" if perms[7] == "r" else "-",
                "w" if perms[8] == "w" else "-",
                "x" if perms[9] in "xt" else "-"
            )
            self.tree.insert("", "end", values=(r["name"], r["type"], r["owner"], r["group"], perm_display))
        self.btn_audit.config(state="normal")
        self.btn_export.config(state="normal")
        self._log("Found %d items in %s\n" % (len(rows), path))

    def _export(self):
        if not self.audit_rows:
            messagebox.showwarning("No Data", "Run a scan first.")
            return
        if not XLSX_OK:
            messagebox.showerror("Missing Library", "openpyxl not installed.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="permissions.xlsx")
        if not path:
            return
        try:
            meta = {"host": self.e_host.get(), "path": self.e_path.get(), "date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}
            export_excel(self.audit_rows, meta, path)
            self._log("Exported to %s\n" % path)
            messagebox.showinfo("Success", "Exported to:\n%s" % path)
        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

    def _get_selected_path(self):
        sel = self.tree.selection()
        if not sel:
            return self.e_path.get()
        item = self.tree.item(sel[0])
        return os.path.join(self.e_path.get(), item["values"][0])

    def _build_chmod(self):
        mode = ""
        # Owner
        mode += "r" if self.chk_or.get() else "-"
        mode += "w" if self.chk_ow.get() else "-"
        mode += "x" if self.chk_ox.get() else "-"
        # Group
        mode += "r" if self.chk_gr.get() else "-"
        mode += "w" if self.chk_gw.get() else "-"
        mode += "x" if self.chk_gx.get() else "-"
        # Others
        mode += "r" if self.chk_otr.get() else "-"
        mode += "w" if self.chk_otw.get() else "-"
        mode += "x" if self.chk_otx.get() else "-"
        # Convert to octal
        octal = ""
        for i in range(3):
            triplet = mode[i*3:(i+1)*3]
            val = 0
            if triplet[0] == "r": val += 4
            if triplet[1] == "w": val += 2
            if triplet[2] == "x": val += 1
            octal += str(val)
        # Special bits
        special = 0
        if self.chk_suid.get(): special += 4
        if self.chk_sgid.get(): special += 2
        if self.chk_sticky.get(): special += 1
        return str(special) + octal if special > 0 else octal

    def _preview(self):
        target = self._get_selected_path()
        owner = self.e_owner.get().strip()
        group = self.e_group.get().strip()
        chmod = self._build_chmod()
        acl_user = self.e_acl_user.get().strip()
        recursive = self.chk_apply_recursive.get()
        
        changes = []
        if owner:
            changes.append("• Change owner to: %s" % owner)
        if group:
            changes.append("• Change group to: %s" % group)
        changes.append("• Set permissions to: %s" % chmod)
        if acl_user:
            acl_perms = ""
            if self.chk_acl_r.get(): acl_perms += "r"
            if self.chk_acl_w.get(): acl_perms += "w"
            if self.chk_acl_x.get(): acl_perms += "x"
            if acl_perms:
                changes.append("• Grant %s access to user: %s" % (acl_perms, acl_user))
        if recursive:
            changes.append("• Apply to all subfolders and files")
        
        msg = "Changes for: %s\n\n%s" % (target, "\n".join(changes))
        messagebox.showinfo("Preview", msg)

    def _apply(self):
        target = self._get_selected_path()
        owner = self.e_owner.get().strip()
        group = self.e_group.get().strip()
        chmod = self._build_chmod()
        acl_user = self.e_acl_user.get().strip()
        recursive = self.chk_apply_recursive.get()
        
        if not messagebox.askyesno("Confirm", "Apply changes to %s?" % target):
            return
        
        self.btn_apply.config(state="disabled")
        self._log("Applying changes...\n")
        threading.Thread(target=self._do_apply, args=(target, owner, group, chmod, acl_user, recursive), daemon=True).start()

    def _do_apply(self, target, owner, group, chmod, acl_user, recursive):
        rflag = "-R" if recursive else ""
        results = []
        
        if owner:
            out, err, code = self.ssh.run("chown %s %s %s" % (rflag, repr(owner), repr(target)))
            results.append(("owner", code == 0))
        
        if group:
            out, err, code = self.ssh.run("chgrp %s %s %s" % (rflag, repr(group), repr(target)))
            results.append(("group", code == 0))
        
        out, err, code = self.ssh.run("chmod %s %s %s" % (rflag, chmod, repr(target)))
        results.append(("permissions", code == 0))
        
        if acl_user:
            acl_perms = ""
            if self.chk_acl_r.get(): acl_perms += "r"
            if self.chk_acl_w.get(): acl_perms += "w"
            if self.chk_acl_x.get(): acl_perms += "x"
            if acl_perms:
                out, err, code = self.ssh.run("setfacl %s -m u:%s:%s %s" % (rflag, acl_user, acl_perms, repr(target)))
                results.append(("ACL", code == 0))
        
        success = sum(1 for _, ok in results if ok)
        self.after(0, lambda: self._apply_done(success, len(results)))

    def _apply_done(self, success, total):
        self.btn_apply.config(state="normal")
        self._log("Applied %d/%d changes successfully.\n" % (success, total))
        if success == total:
            messagebox.showinfo("Success", "All changes applied!")
        else:
            messagebox.showwarning("Partial Success", "Applied %d of %d changes." % (success, total))
        # Refresh
        self._audit()

if __name__ == "__main__":
    App().mainloop()
